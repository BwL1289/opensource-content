import inspect
from dataclasses import dataclass, field
from pathlib import Path
from types import FrameType
from typing import Any, cast

from aws_lambda_powertools import Logger
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from ..utils.constants import Constants

logger = Logger(service="equity_equation.py")

# TODOs:
# 1. Use pydantic instead
# 2. Add formatting to cells (%, numbers, etc.)


@dataclass(frozen=False)
class EquityEquation:
    total_equity_pool: float | int
    _total_equity_pool: float | int = field(
        init=False, repr=False, metadata={"description": "The total equity pool", "example": "100%"}
    )
    proposed_equity_alloc: float | int
    _proposed_equity_alloc: float | int = field(
        init=False,
        repr=False,
        metadata={"description": "The proposed equity allocation to the investor", "example": "7%"},
    )
    estimated_valuation_increase: float | int
    _estimated_valuation_increase: float | int = field(
        init=False,
        repr=False,
        metadata={
            "description": "The estimated increase in company value as a result of the investor investing",
            "example": "7.5%",
        },
    )
    pre_money_valuation: float | int
    _pre_money_valuation: float | int = field(
        init=False,
        repr=False,
        metadata={"description": "The company valuation before the investment", "example": "20,000,000"},
    )

    def __post_init__(self) -> None:
        # Calculate the available equity pool after the investment
        self.available_equity_pool_after_investment: float = self.total_equity_pool - self.proposed_equity_alloc

    @property
    def proposed_equity_alloc(self) -> float:
        current_frame = cast(FrameType, inspect.currentframe())
        logger.info(
            f"""Getting the {current_frame.f_code.co_name}\
                from {cast(FrameType, current_frame.f_back).f_code.co_name}"""
        )
        return self._proposed_equity_alloc

    @proposed_equity_alloc.setter
    def proposed_equity_alloc(self, proposed_equity_alloc: Any) -> None:
        logger.info(f"Setting the proposed_equity_alloc: {proposed_equity_alloc}")

        if not isinstance(proposed_equity_alloc, float | int):
            msg = f"""This proposed_equity_alloc\
                is not a float or int: {type(proposed_equity_alloc)}: {proposed_equity_alloc}"""
            logger.error(msg)
            raise Exception(msg)

        if proposed_equity_alloc <= Constants.ZERO.value or proposed_equity_alloc >= Constants.ONE_HUNDRED.value:
            msg = f"""Invalid value for equity percentage.\
                It should be greater than 0 and less than 100: {proposed_equity_alloc}"""
            logger.error(msg)
            raise Exception(msg)

        self._proposed_equity_alloc = proposed_equity_alloc

    @property
    def estimated_valuation_increase(self) -> float:
        current_frame = cast(FrameType, inspect.currentframe())
        logger.info(
            f"""Getting the {current_frame.f_code.co_name} from\
                {cast(FrameType, current_frame.f_back).f_code.co_name}"""
        )
        return self._estimated_valuation_increase

    @estimated_valuation_increase.setter
    def estimated_valuation_increase(self, estimated_valuation_increase: Any) -> None:
        logger.info(f"Setting the estimated_valuation_increase: {estimated_valuation_increase}")

        if not isinstance(estimated_valuation_increase, float | int):
            msg = f"This estimated_valuation_increase\
                is not a float or int: {type(estimated_valuation_increase)}: {estimated_valuation_increase}"
            logger.error(msg)
            raise Exception(msg)

        if estimated_valuation_increase <= Constants.ZERO.value:
            msg = f"""Invalid value for equity percentage.\
                It should be greater than 0 and less than 100:{estimated_valuation_increase}"""
            logger.error(msg)
            raise Exception(msg)

        self._estimated_valuation_increase = estimated_valuation_increase

    @property
    def total_equity_pool(self) -> float:
        current_frame = cast(FrameType, inspect.currentframe())
        logger.info(
            f"""Getting the {current_frame.f_code.co_name} from\
                {cast(FrameType, current_frame.f_back).f_code.co_name}"""
        )
        return self._total_equity_pool

    @total_equity_pool.setter
    def total_equity_pool(self, total_equity_pool: Any) -> None:
        logger.info(f"Setting the total_equity_pool: {total_equity_pool}")

        if not isinstance(total_equity_pool, float | int):
            msg = f"This total_equity_pool \
                is not a float or int: {type(total_equity_pool)}: {total_equity_pool}"
            logger.error(msg)
            raise Exception(msg)

        if total_equity_pool <= Constants.ZERO.value or total_equity_pool > Constants.ONE_HUNDRED.value:
            msg = f"""Invalid value for equity percentage. \
                It should be greater than 0 and less than 100: {total_equity_pool}"""
            logger.error(msg)
            raise Exception(msg)

        self._total_equity_pool = total_equity_pool

    @property
    def pre_money_valuation(self) -> float:
        current_frame = cast(FrameType, inspect.currentframe())
        logger.info(
            f"""Getting the {current_frame.f_code.co_name} from\
                {cast(FrameType, current_frame.f_back).f_code.co_name}"""
        )
        return self._pre_money_valuation

    @pre_money_valuation.setter
    def pre_money_valuation(self, pre_money_valuation: Any) -> None:
        logger.info(f"Setting the pre_money_valuation: {pre_money_valuation}")

        if not isinstance(pre_money_valuation, float | int):
            msg = f"This pre_money_valuation \
                is not a float or int: {type(pre_money_valuation)}: {pre_money_valuation}"
            logger.error(msg)
            raise Exception(msg)

        if pre_money_valuation <= Constants.ZERO.value:
            msg = f"""Invalid value for equity percentage. \
                It should be greater than 0: {pre_money_valuation}"""
            logger.error(msg)
            raise Exception(msg)

        self._pre_money_valuation = pre_money_valuation

    def _convert_to_decimal(self, value: float | int) -> float | int:
        return value / Constants.ONE_HUNDRED.value

    def _convert_to_percentage(self, value: float | int) -> float | int:
        return value * Constants.ONE_HUNDRED.value

    def _setup_excel_workbook(self) -> tuple[Worksheet, Workbook]:
        # Create a new workbook and select the active worksheet or create a new one
        wb = Workbook(write_only=False)

        ws = wb.active if wb.active is not None else wb.create_sheet()

        # Set the title of the worksheet
        ws.title = "Investor Equity Equation"

        # Freeze the first row
        ws.freeze_panes = "A1"

        # Add headers to the worksheet
        ws.append(
            [
                "Investor Equity Calculation Parameters",
                "Investor Equity Calculation Values",
                "Accept?",
            ]
        )

        ws["A2"] = "Total Equity Pool"
        ws["B2"] = self.total_equity_pool

        ws["A3"] = "Proposed Equity Allocation to Investor"
        ws["B3"] = self.proposed_equity_alloc

        ws["A4"] = "Remaining Equity Pool After Investment"
        ws["B4"] = self.available_equity_pool_after_investment

        ws["A5"] = "Estimated Valuation Increase as a Result of Investment"
        ws["B5"] = self.estimated_valuation_increase

        wb.save(filename="equity_equation.xlsx")

        return ws, wb

    def _calc_req_investor_value(self) -> float:
        """Calculate what the value of the investor return needs to be in order to justify the
        investment."""

        proposed_equity_alloc_decimal = self._convert_to_decimal(self.proposed_equity_alloc)

        try:
            # Calculate the decision value based on the proposed equity allocation (i.e. 1.075)
            decision_value_decimal = Constants.ONE.value / (Constants.ONE.value - proposed_equity_alloc_decimal)
        except ZeroDivisionError:
            raise ZeroDivisionError(
                "Equity percentage cannot be 100. Please enter a value greater than 0 and less than 100%."
            )
        else:
            logger.info(f"The required investor value to justify the investment is: {decision_value_decimal}")
            return decision_value_decimal

    def _calc_est_company_value(self) -> int | float:
        """Calculate the estimated increase/decrease in company value after the investment.

        This will dictate whether to accept or reject an investment or hiring decision based on the
        proposed equity percentage to be allocated to the investor, and the expected return on investment.
        """
        # Convert the estimated valuation increase as a decimal
        estimated_valuation_increase_decimal = self._convert_to_decimal(self.estimated_valuation_increase)

        # Calculate the estimated company value increase/decrease after the investment
        company_value_after_investment = self.pre_money_valuation * (
            Constants.ONE.value + estimated_valuation_increase_decimal
        )
        logger.info(
            f"The estimated company value inc/dec as a percent after the investment is: {company_value_after_investment}"
        )

        return company_value_after_investment

    @staticmethod
    def _get_root_dir() -> str:
        """Get the project's root directory"""
        last_idx = -1
        cur_dir, src = Path(__file__).parent, "src"
        root_dir = str(next(p for p in cur_dir.parents if p.parts[last_idx] == src).parent)

        return root_dir

    def __call__(self) -> None:
        ws, wb = self._setup_excel_workbook()
        decision_value = self._calc_req_investor_value()

        # Convert decision value to a percentage
        decision_value_percentage = self._convert_to_percentage(decision_value - Constants.ONE.value)

        ws["A6"] = "Percentage Increase in Company Value Required to Justify Investment"
        ws["B6"] = round(decision_value_percentage, ndigits=2)

        estimated_new_company_value = self._calc_est_company_value()
        logger.info(f"The estimated new company value is: {estimated_new_company_value}")

        ws["A7"] = "Premoney Valuation"
        ws["B7"] = self.pre_money_valuation

        ws["A8"] = "Estimated Increase/Decrease in Company Value After Investment (not including new money in)"
        ws["B8"] = estimated_new_company_value

        accept_or_reject_value = (
            Constants.ACCEPT.value
            if decision_value_percentage >= self.proposed_equity_alloc
            else Constants.REJECT.value
        )
        logger.info(f"Based on the decision value, the decision is to: {accept_or_reject_value}")
        c2 = ws["C2"]

        # Set the fill color (e.g., green for accept, red for reject)
        cell_color = "008000" if accept_or_reject_value == "Accept" else "FF0000"
        c2.fill = PatternFill(start_color=cell_color, fill_type="solid")

        # Set the font to bold
        c2.font = Font(bold=True, color="FFFFFF")
        c2 = accept_or_reject_value

        wb.save(filename=f"{self._get_root_dir()}/equity_equation.xlsx")
